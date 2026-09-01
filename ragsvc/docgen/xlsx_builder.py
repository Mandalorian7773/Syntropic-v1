"""Excel deliverable generation. Owner: person 2.

An .xlsx that is a grid of left-aligned strings is a CSV wearing a costume.
What makes a workbook usable to the person who receives it is that the numbers
are numbers -- sortable, summable, right-aligned, formatted -- so the values
here are coerced to int and float where they parse, and left as text where they
do not.

The rest is the ordinary courtesy of a spreadsheet somebody else has to work
in: a frozen header, an autofilter, column widths that fit the content, and the
same generated-draft notice the Word documents carry.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import ragconfig as cfg

from .schema import Sheet

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
TITLE_FILL = PatternFill("solid", fgColor="D9E2F3")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=12, color="1F3864")
ORG_FONT = Font(bold=True, size=11, color="1F3864")
NOTICE_FONT = Font(bold=True, size=8, color="C00000")
BODY_FONT = Font(size=10)

THIN = Side(style="thin", color="BFBFBF")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MAX_WIDTH = 60
MIN_WIDTH = 9
# Excel rejects : \ / ? * [ ] in a sheet name and caps it at 31 characters.
INVALID_SHEET_CHARS = re.compile(r"[:\\/?*\[\]]")
NUMERIC = re.compile(r"^-?\d{1,3}(?:,\d{3})*(?:\.\d+)?$|^-?\d+(?:\.\d+)?$")


def _safe_sheet_name(name: str, used: set[str]) -> str:
    cleaned = INVALID_SHEET_CHARS.sub("-", (name or "Sheet").strip()) or "Sheet"
    cleaned = cleaned[:31]
    candidate, suffix = cleaned, 2
    while candidate.lower() in used:
        tail = f"-{suffix}"
        candidate = cleaned[: 31 - len(tail)] + tail
        suffix += 1
    used.add(candidate.lower())
    return candidate


def _coerce(value: Any) -> Any:
    """Turn a numeric-looking string into a number, leave everything else alone.

    The agent emits every cell as a string. A column of "12.5" that Excel
    treats as text cannot be summed, sorted or charted, which defeats the point
    of producing a workbook rather than a table in a chat window.
    """
    if value is None:
        return ""
    if isinstance(value, (int, float, bool)):
        return value
    text = str(value).strip()
    if not text or not NUMERIC.match(text):
        return text
    plain = text.replace(",", "")
    try:
        number = float(plain)
    except ValueError:
        return text
    return int(number) if number.is_integer() and "." not in plain else number


def _write_sheet(worksheet, sheet: Sheet) -> None:
    row_index = 1
    columns = list(sheet.columns)
    rows = [list(r) for r in sheet.rows]

    if not columns and rows:
        columns, rows = [str(c) for c in rows[0]], rows[1:]
    width = max([len(columns)] + [len(r) for r in rows]) if rows else len(columns)
    width = max(width, 1)

    # Letterhead line, so a printed sheet is attributable on paper.
    cell = worksheet.cell(row=row_index, column=1, value=cfg.ORG_NAME)
    cell.font = ORG_FONT
    worksheet.merge_cells(
        start_row=row_index, start_column=1, end_row=row_index, end_column=width
    )
    row_index += 1

    if sheet.title:
        title_cell = worksheet.cell(row=row_index, column=1, value=sheet.title)
        title_cell.font = TITLE_FONT
        title_cell.fill = TITLE_FILL
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        worksheet.merge_cells(
            start_row=row_index, start_column=1, end_row=row_index, end_column=width
        )
        worksheet.row_dimensions[row_index].height = 20
        row_index += 1

    row_index += 1  # a blank spacer row before the table
    header_row = row_index

    for index in range(width):
        name = columns[index] if index < len(columns) else ""
        cell = worksheet.cell(row=header_row, column=index + 1, value=name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = CELL_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    worksheet.row_dimensions[header_row].height = 26

    widths = [max(MIN_WIDTH, min(len(str(c)) + 4, MAX_WIDTH)) for c in columns]
    widths += [MIN_WIDTH] * (width - len(widths))

    for offset, row in enumerate(rows, start=1):
        for index in range(width):
            raw = row[index] if index < len(row) else ""
            value = _coerce(raw)
            cell = worksheet.cell(row=header_row + offset, column=index + 1, value=value)
            cell.font = BODY_FONT
            cell.border = CELL_BORDER
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = "#,##0.00" if isinstance(value, float) else "#,##0"
            else:
                cell.alignment = Alignment(vertical="top", wrap_text=len(str(value)) > 40)
            widths[index] = max(widths[index], min(len(str(value)) + 3, MAX_WIDTH))

    for index in range(width):
        worksheet.column_dimensions[get_column_letter(index + 1)].width = widths[index]

    # Freeze below the header and filter on it, so a hundred-row findings list
    # is workable rather than merely present.
    worksheet.freeze_panes = worksheet.cell(row=header_row + 1, column=1)
    if rows:
        worksheet.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(width)}{header_row + len(rows)}"
        )

    notes_row = header_row + len(rows) + 2
    for note in sheet.notes:
        note_cell = worksheet.cell(row=notes_row, column=1, value=str(note))
        note_cell.font = Font(size=9, italic=True)
        notes_row += 1

    notice = worksheet.cell(row=notes_row + 1, column=1, value=cfg.GENERATED_NOTICE)
    notice.font = NOTICE_FONT

    worksheet.page_setup.orientation = "landscape" if width > 6 else "portrait"
    worksheet.print_title_rows = f"{header_row}:{header_row}"


def build_xlsx(sheets: list[Sheet], path: Path) -> int:
    """Write a workbook to `path`. Returns the number of sheets written."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    used: set[str] = set()

    for sheet in sheets or [Sheet(name="Sheet1")]:
        worksheet = workbook.create_sheet(_safe_sheet_name(sheet.name, used))
        _write_sheet(worksheet, sheet)

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(str(path))
    return len(workbook.worksheets)
